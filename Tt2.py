import asyncio
import re
import subprocess
import sys

import aiohttp
import config
import nest_asyncio
import os
import pytz
import threading
import json
import pandas as pd
import requests
import logging

import telegram.error
from apscheduler.schedulers.asyncio import AsyncIOScheduler

from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, ChatPermissions, \
    BotCommand, BotCommandScopeDefault, BotCommandScopeChat, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, CallbackContext, \
    ContextTypes
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash
from flask import Flask, render_template, request, redirect, url_for, session, \
    jsonify  # Убрать  SocketIO ----------------------
import hashlib
import json
import hashlib

from aiocron import crontab
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill

from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime

from telegram.ext import Application

import tkinter as tk
from tkinter import scrolledtext, simpledialog, messagebox, filedialog

from PIL import Image, ImageTk, ImageGrab
import io

from tkinter import Menu

from gevent import monkey

nest_asyncio.apply()

global muted_users

scheduler = BackgroundScheduler(timezone="Europe/Kiev")

EXCEL_FILE = "user_data_export.xlsx"

application = None

app = Flask(__name__)

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Ключ для сессий

DATA_FILE = "data.json"
CHATS_FILE = "chats.json"
DEFAULT_AVATAR_URL = "https://img2.freepng.ru/20180327/ziq/avjctv0xo.webp"
DEFAULT_AVATAR_PATH = "default_avatar.png"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}


# -------------------------------------------------------------------------------------------------------------------------------
def load_data2():
    """Загружает данные пользователей из JSON-файла."""
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    print(DATA_FILE)
    return {user["id"]: user for user in data["users"]}, data


def load_chats2():
    try:
        with open(CHATS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def download_default_avatar():
    """Скачивает стандартный аватар и сохраняет локально."""
    if os.path.exists(DEFAULT_AVATAR_PATH):  # Если файл уже скачан, просто используем его
        return Image.open(DEFAULT_AVATAR_PATH)

    try:
        response = requests.get(DEFAULT_AVATAR_URL, timeout=5, stream=True)
        response.raise_for_status()  # Проверяем на ошибки загрузки

        with open(DEFAULT_AVATAR_PATH, "wb") as f:
            f.write(response.content)

        return Image.open(io.BytesIO(response.content))  # Открываем скачанное изображение
    except requests.exceptions.RequestException as e:
        print(f"Ошибка загрузки стандартного аватара: {e}")
        return None  # Если не удалось скачать, возвращаем None


def check_avatar(user_id):
    """Проверяет наличие аватара пользователя и возвращает URL аватара или путь к стандартному аватару."""
    try:
        response = requests.get(f"{TELEGRAM_API_URL}getUserProfilePhotos?user_id={user_id}&limit=1", timeout=5)
        data = response.json()

        if data.get("result", {}).get("photos"):
            file_id = data["result"]["photos"][0][0]["file_id"]
            file_info = requests.get(f"{TELEGRAM_API_URL}getFile?file_id={file_id}", timeout=5).json()
            file_path = file_info["result"]["file_path"]
            return f"https://api.telegram.org/file/bot{BOTTOCEN}/{file_path}"
    except requests.exceptions.RequestException as e:
        print(f"Ошибка загрузки аватара Telegram: {e}")

    return DEFAULT_AVATAR_PATH  # Если нет аватара, возвращаем путь к стандартному аватару


def download_image(url):
    """Пытается загрузить изображение по URL и вернуть объект PIL.Image."""
    try:
        response = requests.get(url, timeout=5, stream=True)
        response.raise_for_status()
        return Image.open(io.BytesIO(response.content))
    except requests.exceptions.RequestException:
        return None


def get_user_avatar(user_id):
    """Загружает аватар пользователя с приоритетом: Telegram -> Ссылка -> Локальный файл."""
    avatar_url = check_avatar(user_id)

    # 1️⃣ Пытаемся загрузить аватар из Telegram
    if avatar_url:
        avatar = download_image(avatar_url)
        if avatar:
            return avatar

    # 2️⃣ Если нет — пробуем загрузить изображение по ссылке
    avatar = download_image(DEFAULT_AVATAR_URL)
    if avatar:
        return avatar

    # 3️⃣ Если и ссылка не работает — используем локальный файл
    if os.path.exists(DEFAULT_AVATAR_PATH):
        return Image.open(DEFAULT_AVATAR_PATH)

    return None


def save_message_to_json(user_id, username, message):
    """Добавляет сообщение в chats.json с флагом прочитанности"""
    chats_data = load_chats()
    chat_id_str = str(user_id)

    new_message = {
        "username": username,
        "message": message,
        "time_sent": datetime.now().strftime("%H:%M; %d/%m/%Y"),
        "read": False  # Добавляем статус "непрочитанное"
    }

    # Если у пользователя еще нет сообщений, создаем список
    if chat_id_str not in chats_data:
        chats_data[chat_id_str] = {"username": username, "messages": []}

    # Добавляем сообщение
    chats_data[chat_id_str]["messages"].append(new_message)

    # Сохраняем изменения в файл
    with open(CHATS_FILE, "w", encoding="utf-8") as file:
        json.dump(chats_data, file, ensure_ascii=False, indent=4)

    return True


def update_second_name(user_id, new_second_name, file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)

        user_found = False
        for user in data['users']:
            if user['id'] == user_id:
                user['second_name'] = new_second_name
                user_found = True
                break

        if user_found:
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(data, file, ensure_ascii=False, indent=4)
            print(f"Имя пользователя с ID {user_id} успешно обновлено.")

        else:
            print(f"Пользователь с ID {user_id} не найден.")

    except Exception as e:
        print(f"Ошибка при обновлении имени пользователя: {e}")


class RoundedFrame(tk.Canvas):
    """Кастомный фрейм с закругленными углами."""

    def __init__(self, master, radius=20, bg="white", **kwargs):
        super().__init__(master, bg=bg, highlightthickness=0, **kwargs)
        self.radius = radius
        self.bg = bg
        self.bind("<Configure>", self._draw_rounded_rect)

    def _draw_rounded_rect(self, event=None):
        self.delete("all")
        width = self.winfo_width()
        height = self.winfo_height()
        self.create_rounded_rect(0, 0, width, height, radius=self.radius, fill=self.bg)

    def create_rounded_rect(self, x1, y1, x2, y2, radius=20, **kwargs):
        points = [
            x1 + radius, y1,
            x2 - radius, y1,
            x2, y1,
            x2, y1 + radius,
            x2, y2 - radius,
            x2, y2,
            x2 - radius, y2,
            x1 + radius, y2,
            x1, y2,
            x1, y2 - radius,
            x1, y1 + radius,
            x1, y1,
        ]
        return self.create_polygon(points, **kwargs, smooth=True)


class ChatApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Пользователи и чат")

        self.users, self.data = load_data2()
        self.chats = load_chats2()
        self.file_path = 'data.json'
        self.bot_token = BOTTOCEN  # Замените на ваш токен бота

        self.main_frame = tk.Frame(root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Фрейм для списка пользователей с прокруткой
        self.user_list_frame = tk.Frame(self.main_frame)
        self.user_list_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)

        # Canvas и Scrollbar для списка пользователей
        self.user_canvas = tk.Canvas(self.user_list_frame)
        self.user_scrollbar = tk.Scrollbar(self.user_list_frame, orient=tk.VERTICAL, command=self.user_canvas.yview)
        self.user_list_container = tk.Frame(self.user_canvas)

        # Привязка контейнера к Canvas
        self.user_list_container.bind(
            "<Configure>",
            lambda e: self.user_canvas.configure(scrollregion=self.user_canvas.bbox("all")))
        self.user_canvas.create_window((0, 0), window=self.user_list_container, anchor="nw")
        self.user_canvas.configure(yscrollcommand=self.user_scrollbar.set)

        # Привязка колесика мыши к Canvas
        self.user_canvas.bind_all("<MouseWheel>", self.on_mousewheel)

        # Размещение Canvas и Scrollbar
        self.user_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.user_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.user_buttons = {}

        # Добавление пользователей в контейнер
        for user_id, user in self.users.items():
            self.create_user_block(user_id, user)

        # Фрейм для чата
        self.chat_frame = tk.Frame(self.main_frame)
        self.chat_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Фрейм для отображения аватара и информации о пользователе
        self.header_frame = tk.Frame(self.chat_frame, bd=1, relief=tk.SOLID, padx=5, pady=5)
        self.header_frame.pack(fill=tk.X, padx=5, pady=5)

        # Аватар пользователя
        self.avatar_label = tk.Label(self.header_frame, bd=2, relief="solid")
        self.avatar_label.pack(side=tk.LEFT, padx=5)

        # Информация о пользователе
        self.user_info_frame = tk.Frame(self.header_frame)
        self.user_info_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.second_name_label = tk.Label(self.user_info_frame, font=("Helvetica", 12, "bold"), anchor="w")
        self.second_name_label.pack(fill=tk.X)

        self.username_label = tk.Label(self.user_info_frame, font=("Helvetica", 10), anchor="w")
        self.username_label.pack(fill=tk.X)

        self.user_id_label = tk.Label(self.user_info_frame, font=("Helvetica", 10), anchor="w")
        self.user_id_label.pack(fill=tk.X)

        # Кнопка Замутить/Размутить
        self.mute_button = tk.Button(self.header_frame, text="Замутить", command=self.toggle_mute)
        self.mute_button.pack(side=tk.RIGHT, padx=5)

        # Чат
        self.chat_canvas = tk.Canvas(self.chat_frame, bg="#f0f0f0", highlightthickness=0)
        self.chat_scrollbar = tk.Scrollbar(self.chat_frame, orient=tk.VERTICAL, command=self.chat_canvas.yview)
        self.chat_container = tk.Frame(self.chat_canvas, bg="#f0f0f0")

        # Привязка контейнера к Canvas
        self.chat_container.bind(
            "<Configure>",
            lambda e: self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all")))
        self.chat_canvas.create_window((0, 0), window=self.chat_container, anchor="nw")
        self.chat_canvas.configure(yscrollcommand=self.chat_scrollbar.set)

        # Привязка колесика мыши к Canvas чата
        self.chat_canvas.bind_all("<MouseWheel>", self.on_mousewheel_chat)

        # Размещение Canvas и Scrollbar
        self.chat_canvas.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.chat_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Фрейм для превью файлов (скрыт по умолчанию)
        self.file_preview_frame = tk.Frame(self.chat_frame, bg="#f0f0f0")

        # Canvas и Scrollbar для превью файлов
        self.file_preview_canvas = tk.Canvas(self.file_preview_frame, bg="#f0f0f0", highlightthickness=0)
        self.file_preview_scrollbar = tk.Scrollbar(self.file_preview_frame, orient=tk.VERTICAL,
                                                   command=self.file_preview_canvas.yview)
        self.file_preview_container = tk.Frame(self.file_preview_canvas, bg="#f0f0f0")

        # Привязка контейнера к Canvas
        self.file_preview_container.bind(
            "<Configure>",
            lambda e: self.file_preview_canvas.configure(scrollregion=self.file_preview_canvas.bbox("all")))
        self.file_preview_canvas.create_window((0, 0), window=self.file_preview_container, anchor="nw")
        self.file_preview_canvas.configure(yscrollcommand=self.file_preview_scrollbar.set)

        # Привязка колесика мыши к Canvas
        self.file_preview_canvas.bind_all("<MouseWheel>", self.on_mousewheel_file_preview)

        # Размещение Canvas и Scrollbar
        self.file_preview_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.file_preview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Поле ввода сообщения и кнопка отправки (внизу, под чатом)
        self.entry_frame = tk.Frame(self.chat_frame, bg="#f0f0f0")
        self.entry_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=5)

        # Кнопка для переключения видимости контейнера с превью
        self.toggle_preview_button = tk.Button(
            self.entry_frame,
            text="📁",  # Иконка папки
            font=("Helvetica", 12),
            command=self.toggle_file_preview,  # Метод для переключения видимости
        )
        self.toggle_preview_button.pack(side=tk.LEFT, padx=5)

        # Кнопка для отправки документов
        self.send_document_button = tk.Button(
            self.entry_frame,
            text="📄",  # Иконка документа
            font=("Helvetica", 12),
            command=self.send_document,  # Метод для отправки документа
        )
        self.send_document_button.pack(side=tk.LEFT, padx=5)

        self.chat_input = tk.Text(self.entry_frame, font=("Helvetica", 12), height=3)
        self.chat_input.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.chat_input.bind("<KeyPress>", self.key_press_handler)  # Привязываем обработчик нажатий клавиш
        self.chat_input.bind("<Control-v>", self.paste_image)  # Привязываем вставку картинки
        self.send_button = tk.Button(self.entry_frame, text="Отправить", command=self.send_message,
                                     font=("Helvetica", 12))
        self.send_button.pack(side=tk.RIGHT)

        self.current_user_id = None
        self.mute_end_label = None

    def create_user_block(self, user_id, user):
        """Создает блок пользователя в списке."""
        user_frame = tk.Frame(self.user_list_container, bd=1, relief=tk.SOLID, padx=5, pady=5)
        user_frame.pack(fill=tk.X, padx=5, pady=2)

        # Аватар пользователя
        avatar = get_user_avatar(user_id)
        if avatar:
            avatar = avatar.resize((40, 40))  # Возвращаем старый размер аватара
            avatar_image = ImageTk.PhotoImage(avatar)
            avatar_label = tk.Label(user_frame, image=avatar_image, bd=2, relief="solid")
            avatar_label.image = avatar_image
            avatar_label.pack(side=tk.LEFT, padx=5)
            if user.get("mute", False):
                avatar_label.config(highlightbackground="red", highlightcolor="red", highlightthickness=2)

        # Информация о пользователе
        user_label = tk.Label(
            user_frame,
            text=f"{user['second_name']} ({user['username']})",
            font=("Helvetica", 12, "bold"),  # Возвращаем старый шрифт
            anchor="w",
            cursor="hand2",
        )
        user_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        user_label.bind("<Button-1>", lambda event, uid=user_id: self.open_chat(uid))

        # Кнопка редактирования
        edit_button = tk.Button(
            user_frame,
            text="✏️",
            font=("Helvetica", 12),  # Возвращаем старый шрифт
            command=lambda uid=user_id: self.edit_user_name(uid),
        )
        edit_button.pack(side=tk.RIGHT, padx=5)

        # Количество непрочитанных сообщений (справа снизу, поверх аватара)
        unread_count = self.get_unread_message_count(user_id)
        if unread_count > 0:
            unread_label = tk.Label(
                user_frame,
                text=f"{unread_count}",
                font=("Helvetica", 10, "bold"),
                bg="red",
                fg="white",
                bd=2,
                relief="solid",
            )
            unread_label.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")  # Позиционируем справа снизу

        self.user_buttons[user_id] = user_frame

    def bind_mousewheel(self, widget, handler):
        """Рекурсивно привязывает событие прокрутки ко всем дочерним элементам."""
        widget.bind("<MouseWheel>", handler)
        for child in widget.winfo_children():
            self.bind_mousewheel(child, handler)

    def on_mousewheel(self, event):
        """Обработчик события прокрутки колесика мыши для списка пользователей."""
        if self.user_canvas.bbox("all"):  # Проверяем, есть ли что прокручивать
            self.user_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def on_mousewheel_chat(self, event):
        """Обработчик события прокрутки колесика мыши для чата."""
        if self.chat_canvas.bbox("all"):  # Проверяем, есть ли что прокручивать
            self.chat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def update_message_width(self, message_text, height):
        """Обновляет ширину текстового поля в зависимости от ширины окна."""
        window_width = self.root.winfo_width()
        message_width = int((window_width - 550) / 10)  # Уменьшаем длину сообщения
        message_text.config(width=message_width)

        # Обновляем высоту текстового поля в зависимости от количества строк
        message_text_height = int(len(message_text.get("1.0", tk.END))) / int((window_width - 550) / 10) + height
        message_text.config(height=message_text_height)

    def update_user_info(self, user):
        """Обновляет информацию о пользователе в header_frame."""
        # Загружаем аватар пользователя
        avatar = get_user_avatar(user["id"])
        if avatar:
            avatar = avatar.resize((50, 50))  # Размер аватара
            avatar_image = ImageTk.PhotoImage(avatar)
            self.avatar_label.config(image=avatar_image)
            self.avatar_label.image = avatar_image  # Сохраняем ссылку, чтобы избежать сборки мусора
        else:
            self.avatar_label.config(image=None)

        # Обновляем информацию о пользователе
        self.second_name_label.config(text=user["second_name"])
        self.username_label.config(text=f"@{user['username']}")
        self.user_id_label.config(text=f"ID: {user['id']}")

    def paste_image(self, event=None):
        """Вставляет картинку из буфера обмена в поле ввода."""
        try:
            # Получаем картинку из буфера обмена
            image = ImageGrab.grabclipboard()
            if image:
                # Преобразуем картинку в формат, который можно вставить в Text
                image = image.resize((100, 100))  # Уменьшаем размер для удобства
                photo = ImageTk.PhotoImage(image)
                self.chat_input.image_create(tk.END, image=photo)
                self.chat_input.insert(tk.END, "\n")  # Добавляем перенос строки
        except Exception as e:
            print(f"Ошибка при вставке картинки: {e}")

    def send_document(self):
        """Открывает проводник для выбора документа и отображает его перед отправкой."""
        if not self.current_user_id:
            return

        # Открываем проводник для выбора файла
        file_path = filedialog.askopenfilename(
            title="Выберите документ",
            filetypes=[("Все файлы", "*.*")],  # Можно указать конкретные типы файлов
        )

        if file_path:
            self.add_file_to_preview(file_path)

    def send_photo(self):
        """Открывает проводник для выбора картинки и отображает ее перед отправкой."""
        if not self.current_user_id:
            return

        # Открываем проводник для выбора файла
        file_path = filedialog.askopenfilename(
            title="Выберите картинку",
            filetypes=[("Все файлы", "*.*")],  # Только изображения
        )

        if file_path:
            self.add_file_to_preview(file_path, "photo")

    def send_file(self, file_path):
        """Отправляет файл через Telegram Bot API."""
        if not self.current_user_id:
            return

        # Определяем тип файла по его расширению
        file_extension = file_path.split(".")[-1].lower()
        file_type = "document"  # По умолчанию отправляем как документ

        if file_extension in ["jpg", "jpeg", "png", "gif", "bmp"]:
            file_type = "photo"  # Если это изображение, отправляем как фото

        url = f"https://api.telegram.org/bot{self.bot_token}/send{file_type}"
        files = {file_type: open(file_path, "rb")}
        data = {"chat_id": self.current_user_id}

        try:
            response = requests.post(url, files=files, data=data)
            response.raise_for_status()  # Проверка на ошибки HTTP
            print(f"Файл успешно отправлен: {file_path}")
        except requests.exceptions.HTTPError as http_err:
            print(f"Ошибка HTTP при отправке файла: {http_err}")
        except Exception as err:
            print(f"Ошибка при отправке файла: {err}")
        finally:
            files[file_type].close()  # Закрываем файл после отправки

    def send_message(self):
        """Отправляет сообщение и все выбранные файлы."""
        if not self.current_user_id:
            return

        # Отправляем текстовое сообщение
        message = self.chat_input.get("1.0", tk.END).strip()
        if message:
            self.send_telegram_message(self.current_user_id, message)

        # Отправляем все выбранные файлы
        if hasattr(self, "files_to_send"):
            for file in self.files_to_send:
                self.send_file(file["path"])  # Передаем только путь к файлу
            self.files_to_send = []  # Очищаем список файлов после отправки

        # Очищаем поле ввода и превью файлов
        self.chat_input.delete("1.0", tk.END)
        for widget in self.file_preview_container.winfo_children():
            widget.destroy()

    def open_file(self, file_path):
        """Открывает файл в соответствующей программе."""
        try:
            if os.name == "nt":  # Для Windows
                os.startfile(file_path)
            else:  # Для macOS и Linux
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.call([opener, file_path])
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")

    def add_file_to_preview(self, file_path):
        """Добавляет файл в превью перед отправкой."""
        file_frame = tk.Frame(self.file_preview_container, bg="#f0f0f0")
        file_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        # Получаем расширение файла
        file_extension = file_path.split(".")[-1].lower()

        # Определяем иконку в зависимости от расширения файла
        icon_mapping = {
            "jpg": "🖼️",  # Иконка для изображений
            "jpeg": "🖼️",
            "png": "🖼️",
            "gif": "🖼️",
            "bmp": "🖼️",
            "py": "🐍",  # Иконка для Python файлов
            "exe": "⚙️",  # Иконка для исполняемых файлов
            "docx": "📄",  # Иконка для Word документов
            "xlsx": "📊",  # Иконка для Excel файлов
            "txt": "📝",  # Иконка для текстовых файлов
            "pdf": "📑",  # Иконка для PDF файлов
        }

        # Иконка по умолчанию для неизвестных типов файлов
        file_icon = icon_mapping.get(file_extension, "📁")

        # Отображаем иконку файла
        icon_label = tk.Label(file_frame, text=file_icon, font=("Helvetica", 20), bg="#f0f0f0")
        icon_label.pack(side=tk.LEFT, padx=5)

        # Добавляем обработчик событий для открытия файла
        icon_label.bind("<Button-1>", lambda e, path=file_path: self.open_file(path))

        # Отображаем имя файла
        file_name = file_path.split("/")[-1]  # Получаем только имя файла
        file_label = tk.Label(file_frame, text=file_name, bg="#f0f0f0", font=("Helvetica", 10))
        file_label.pack(side=tk.LEFT, padx=5)

        # Кнопка для удаления файла
        delete_button = tk.Button(
            file_frame,
            text="❌",  # Иконка крестика
            font=("Helvetica", 10),
            bg="#f0f0f0",
            command=lambda: self.remove_file_from_preview(file_frame, file_path),
        )
        delete_button.pack(side=tk.RIGHT, padx=5)

        # Сохраняем информацию о файле
        if not hasattr(self, "files_to_send"):
            self.files_to_send = []
        self.files_to_send.append({"path": file_path})

    def toggle_file_preview(self):
        """Переключает видимость контейнера с превью файлов."""
        if self.file_preview_frame.winfo_ismapped():
            self.file_preview_frame.pack_forget()  # Скрываем контейнер
        else:
            self.file_preview_frame.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True, padx=5,
                                         pady=5)  # Показываем контейнер

    def remove_file_from_preview(self, file_frame, file_path):
        """Удаляет файл из превью."""
        file_frame.destroy()
        self.files_to_send = [file for file in self.files_to_send if file["path"] != file_path]

    def on_mousewheel_file_preview(self, event):
        """Обработчик события прокрутки колесика мыши для блока с файлами."""
        if self.file_preview_canvas.bbox("all"):  # Проверяем, есть ли что прокручивать
            self.file_preview_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def send_message_event(self, event):
        """Отправляет сообщение и обновляет чат (для привязки к событию)."""
        self.send_message()
        return "break"  # Прерываем дальнейшую обработку события

    def show_context_menu(self, event):
        """Показывает контекстное меню для копирования текста."""
        context_menu = Menu(self.root, tearoff=0)
        context_menu.add_command(label="Копировать", command=lambda: self.copy_text(event.widget))
        context_menu.tk_popup(event.x_root, event.y_root)

    def copy_text(self, widget):
        """Копирует выделенный текст в буфер обмена."""
        try:
            selected_text = widget.selection_get()
            self.root.clipboard_clear()
            self.root.clipboard_append(selected_text)
            self.root.update()  # Обновляем буфер обмена
        except tk.TclError:
            pass

    def copy_text_event(self, event):
        """Обрабатывает событие копирования текста."""
        widget = self.root.focus_get()
        if isinstance(widget, tk.Text):
            self.copy_text(widget)
            return "break"  # Прерываем дальнейшую обработку события

    def paste_text(self, event):
        """Вставляет текст из буфера обмена в поле ввода."""
        try:
            clipboard_text = self.root.clipboard_get()
            self.chat_input.insert(tk.INSERT, clipboard_text)
            return "break"  # Прерываем дальнейшую обработку события
        except tk.TclError:
            pass

    def paste_text_event(self, widget):
        """Вставляет текст в указанное текстовое поле."""
        try:
            clipboard_text = self.root.clipboard_get()
            widget.insert(tk.INSERT, clipboard_text)
            return "break"  # Прерываем дальнейшую обработку события
        except tk.TclError:
            pass

    def key_press_handler(self, event):
        """Обрабатывает нажатия клавиш для копирования и вставки."""
        if event.state & 0x4:  # Проверяем нажатие Ctrl
            if event.keycode in (86, 118):  # Ctrl+V
                self.paste_text(event)
            elif event.keycode in (67, 99):  # Ctrl+C
                self.copy_text_event(event)

    def save_data2(self):
        """Сохраняет данные в JSON-файл, обновляя только нужные поля."""
        try:
            # Обновляем данные пользователей
            self.data["users"] = list(self.users.values())

            # Убедимся, что muted_users существует в self.data
            if "muted_users" not in self.data:
                self.data["muted_users"] = {}

            # Сохраняем обновленные данные обратно в файл
            with open(self.file_path, "w", encoding="utf-8") as file:
                json.dump(self.data, file, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Ошибка при сохранении данных: {e}")

    def toggle_mute(self):
        """Переключает состояние мута пользователя."""
        if not self.current_user_id:
            return

        user = self.users[self.current_user_id]
        if user.get("mute", False):
            self.unmute_user()
        else:
            self.mute_user()

    def mute_user(self):
        """Открывает окно для настройки мута пользователя."""
        mute_window = tk.Toplevel(self.root)
        mute_window.title("Замутить пользователя")

        tk.Label(mute_window, text="Время (дни часы минуты секунды):").pack(pady=5)
        mute_time_entry = tk.Entry(mute_window)
        mute_time_entry.pack(pady=5)
        mute_time_entry.insert(0, "00 00 05 00")  # Значение по умолчанию 5 минут

        tk.Label(mute_window, text="Причина:").pack(pady=5)
        reason_entry = tk.Entry(mute_window)
        reason_entry.pack(pady=5)
        reason_entry.insert(0, "По рішенню адміністратора")  # Значение по умолчанию

        def confirm_mute():
            mute_time_str = mute_time_entry.get()
            reason = reason_entry.get() or "По рішенню адміністратора"

            # Разбор времени мута
            days, hours, minutes, seconds = map(int, mute_time_str.split())
            mute_time = timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds).total_seconds()

            # Обновление данных пользователя
            self.users[self.current_user_id]["mute"] = True
            self.users[self.current_user_id]["mute_end"] = (datetime.now() + timedelta(seconds=mute_time)).strftime(
                "%H:%M; %d/%m/%Y")
            self.users[self.current_user_id]["reason"] = reason

            # Добавляем username в muted_users
            username = self.users[self.current_user_id]["username"]
            if "muted_users" not in self.data:
                self.data["muted_users"] = {}
            self.data["muted_users"][username] = True  # Добавляем запись "username": true

            self.save_data2()  # Сохраняем изменения

            # Отправка сообщения пользователю о мутах
            self.send_telegram_message(self.users[self.current_user_id]["id"],
                                       f"Вас замутили на {str(timedelta(seconds=mute_time))}\nПричина: {reason}")

            # Закрытие окна и обновление чата
            mute_window.destroy()
            self.update_user_list()
            self.open_chat(self.current_user_id)  # Обновляем чат

        tk.Button(mute_window, text="Подтвердить", command=confirm_mute).pack(pady=10)

    def unmute_user(self):
        """Открывает окно для подтверждения размута пользователя."""
        unmute_window = tk.Toplevel(self.root)
        unmute_window.title("Размутить пользователя")

        tk.Label(unmute_window, text="Вы уверены, что хотите размутить пользователя?").pack(pady=10)

        def confirm_unmute():
            # Обновление данных пользователя
            self.users[self.current_user_id]["mute"] = False
            self.users[self.current_user_id]["mute_end"] = None
            self.users[self.current_user_id]["reason"] = None

            # Удаляем username из muted_users
            username = self.users[self.current_user_id]["username"]
            if "muted_users" in self.data and username in self.data["muted_users"]:
                del self.data["muted_users"][username]  # Удаляем запись полностью

            self.save_data2()  # Сохраняем изменения

            # Отправка сообщения пользователю о размутах
            self.send_telegram_message(self.users[self.current_user_id]["id"], "Вы были размучены.")

            # Закрытие окна и обновление чата
            unmute_window.destroy()
            self.update_user_list()
            self.open_chat(self.current_user_id)  # Обновляем чат

        tk.Button(unmute_window, text="Подтвердить", command=confirm_unmute).pack(pady=10)

    def send_telegram_message(self, user_id, message):
        """Отправляет сообщение пользователю в Telegram"""
        # Здесь должен быть код для отправки сообщения пользователю в Telegram
        # Например, используя библиотеку requests для отправки сообщения через Telegram Bot API
        chat_id = user_id  # Предполагается, что user_id соответствует chat_id в Telegram
        url = f"https://api.telegram.org/bot{BOTTOCEN}/sendMessage"
        payload = {
            "chat_id": chat_id,
            "text": message
        }
        payload = {
            "chat_id": chat_id,
            "text": message
        }
        try:
            response = requests.post(url, json=payload)
            response.raise_for_status()  # Проверка на ошибки HTTP
            print("Сообщение успешно отправлено!")
        except requests.exceptions.HTTPError as http_err:
            print(f"Ошибка HTTP: {http_err}")
        except Exception as err:
            print(f"Ошибка отправки сообщения: {err}")

        print("-=--=-=-=-=-=-=-=-=-=-=---=")
        print(url)

        if response.status_code != 200:
            print(f"Ошибка отправки сообщения: {response.text}")

    def edit_user_name(self, user_id):
        """Открывает окно редактирования имени пользователя и обновляет данные."""
        old_name = self.users[user_id]['second_name']
        new_name = simpledialog.askstring("Изменить имя пользователя", f"Старое имя: {old_name}\nВведите новое имя:",
                                          initialvalue=old_name)

        if new_name and new_name.strip():
            self.users[user_id]['second_name'] = new_name.strip()
            self.update_user_list()
            # Используем новую функцию для обновления только second_name
            update_second_name(user_id, new_name.strip(), self.file_path)
            self.open_chat(self.current_user_id)  # Обновляем чат

    def update_second_name(user_id, new_second_name, file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)

            user_found = False
            for user in data['users']:
                if user['id'] == user_id:
                    user['second_name'] = new_second_name
                    user_found = True
                    break

            if user_found:
                with open(file_path, 'w', encoding='utf-8') as file:
                    json.dump(data, file, ensure_ascii=False, indent=4)
                print(f"Имя пользователя с ID {user_id} успешно обновлено.")
            else:
                print(f"Пользователь с ID {user_id} не найден.")

        except Exception as e:
            print(f"Ошибка при обновлении имени пользователя: {e}")

    def open_chat(self, user_id):
        """Открывает чат с выбранным пользователем."""
        self.current_user_id = user_id
        user = self.users[user_id]

        # Обновляем информацию о пользователе в header_frame
        self.update_user_info(user)

        # Очищаем чат и загружаем сообщения
        for widget in self.chat_container.winfo_children():
            widget.destroy()

        self.chats = load_chats2()

        if user_id in self.chats:
            messages = self.chats[user_id]["messages"]
            current_date = None

            for msg in messages:
                # Получаем дату и время сообщения
                try:
                    message_time = datetime.strptime(msg["time_sent"], "%H:%M; %d/%m/%Y").strftime("%H:%M")
                    message_date = datetime.strptime(msg["time_sent"], "%H:%M; %d/%m/%Y").strftime("%Y-%m-%d")
                except ValueError as e:
                    print(f"Ошибка парсинга времени: {e}")
                    continue  # Пропускаем сообщение с некорректным форматом времени

                # Определяем, кто отправил сообщение
                is_bot = msg["username"] == "SupportBot"

                # Создаем контейнер для сообщения
                message_frame = tk.Frame(self.chat_container, bg="#f0f0f0")
                message_frame.pack(fill=tk.X, padx=5, pady=2)

                # Внутренний фрейм для сообщения (с закругленными углами)
                inner_frame = RoundedFrame(
                    message_frame,
                    radius=15,
                    bg="#e0e0e0" if is_bot else "#d1e7ff",  # Цвет фона для бота и пользователя
                )
                # Выравниваем внутренний фрейм по правому краю с отступом 25 пикселей
                inner_frame.pack(side=tk.RIGHT, padx=(0, 0), pady=2)  # Отступ 25 пикселей от правого края

                # Добавляем текст сообщения
                message_text = tk.Text(
                    inner_frame,
                    wrap=tk.WORD,
                    font=("Helvetica", 12),
                    bg="#e0e0e0" if is_bot else "#d1e7ff",
                    relief=tk.FLAT,
                    height=len(msg["message"].split("\n")),  # Высота зависит от количества строк
                )
                message_text.insert(tk.END, msg["message"])
                message_text.config(state=tk.NORMAL)  # Позволяем выделение и копирование текста
                message_text.bind("<Button-3>", self.show_context_menu)  # Привязываем контекстное меню
                message_text.bind("<Control-Key>", self.key_press_handler)  # Привязываем обработчик нажатий клавиш
                message_text.pack(side=tk.TOP, padx=10, pady=5, anchor="e")

                # Привязка колесика мыши к Canvas
                self.bind_mousewheel(message_text, self.on_mousewheel_chat)

                # Обновляем ширину текстового поля при изменении размера окна
                self.update_message_width(message_text, len(msg["message"].split("\n")))

                # Привязываем событие изменения размера окна к обновлению ширины текстового поля
                self.root.bind("<Configure>", lambda event, mt=message_text: self.update_message_width(mt, len(
                    msg["message"].split("\n"))))

                # Добавляем время отправки (внутри контейнера сообщения)
                time_label = tk.Label(
                    inner_frame,
                    text=message_time,
                    font=("Helvetica", 10, "italic"),
                    bg="#e0e0e0" if is_bot else "#d1e7ff",
                    fg="green",
                )
                time_label.pack(side=tk.RIGHT, padx=10, pady=(0, 5),
                                anchor="se")  # Выравниваем время по правому нижнему углу

                # Если дата изменилась, добавляем метку с датой (после сообщения пользователя)
                if not is_bot and message_date != current_date:
                    current_date = message_date
                    date_label = tk.Label(
                        self.chat_container,
                        text=current_date,
                        font=("Helvetica", 10, "bold"),
                        bg="#f0f0f0",
                        fg="gray",
                    )
                    date_label.pack(fill=tk.X, pady=(10, 5), anchor="w")

        # Обновляем область прокрутки
        self.chat_canvas.configure(scrollregion=self.chat_canvas.bbox("all"))

        # Устанавливаем фокус на поле ввода сообщения
        self.chat_input.focus_set()

        # Обновляем состояние кнопки Замутить/Размутить
        self.update_mute_button()

        # Обновляем цвет рамки аватара в header_frame
        self.update_avatar_border()

        self.mark_all_messages_as_read(user_id)
        self.update_user_list()

    def update_avatar_border(self):
        """Обновляет цвет рамки аватара в зависимости от состояния мута."""
        if self.current_user_id and self.users[self.current_user_id].get("mute", False):
            self.avatar_label.config(highlightbackground="red", highlightcolor="red", highlightthickness=2)
        else:
            self.avatar_label.config(highlightbackground="black", highlightcolor="black", highlightthickness=2)

    def update_mute_button(self):
        """Обновляет текст кнопки Замутить/Размутить."""
        if self.current_user_id and self.users[self.current_user_id].get("mute", False):
            self.mute_button.config(text="Размутить")
        else:
            self.mute_button.config(text="Замутить")

    def mark_all_messages_as_read(self, user_id):
        chats_data = load_chats2()
        chat_id_str = str(user_id)

        if chat_id_str in chats_data:
            for msg in chats_data[chat_id_str]["messages"]:
                msg["read"] = True
            chats_data[chat_id_str]["newmassages"] = 0

            with open(CHATS_FILE, "w", encoding="utf-8") as file:
                json.dump(chats_data, file, ensure_ascii=False, indent=4)

    def get_unread_message_count(self, user_id):
        chats_data = load_chats2()
        chat_id_str = str(user_id)
        unread_count = 0

        if chat_id_str in chats_data:
            for msg in chats_data[chat_id_str]["messages"]:
                try:
                    if not msg["read"]:
                        unread_count += 1
                except:
                    print(msg)
        return unread_count

    def update_user_list(self):
        """Обновляет список пользователей."""
        for user_id, user_frame in self.user_buttons.items():
            # Очищаем фрейм пользователя
            for widget in user_frame.winfo_children():
                widget.destroy()

            # Аватар пользователя
            avatar = get_user_avatar(user_id)
            if avatar:
                avatar = avatar.resize((40, 40))  # Возвращаем старый размер аватара
                avatar_image = ImageTk.PhotoImage(avatar)
                avatar_label = tk.Label(user_frame, image=avatar_image, bd=2, relief="solid")
                avatar_label.image = avatar_image
                avatar_label.pack(side=tk.LEFT, padx=5)
                if self.users[user_id].get("mute", False):
                    avatar_label.config(highlightbackground="red", highlightcolor="red", highlightthickness=2)

            # Информация о пользователе
            user_label = tk.Label(
                user_frame,
                text=f"{self.users[user_id]['second_name']} ({self.users[user_id]['username']})",
                font=("Helvetica", 12, "bold"),  # Возвращаем старый шрифт
                anchor="w",
                cursor="hand2",
            )
            user_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
            user_label.bind("<Button-1>", lambda event, uid=user_id: self.open_chat(uid))

            # Кнопка редактирования
            edit_button = tk.Button(
                user_frame,
                text="✏️",
                font=("Helvetica", 12),  # Возвращаем старый шрифт
                command=lambda uid=user_id: self.edit_user_name(uid),
            )
            edit_button.pack(side=tk.RIGHT, padx=5)

            # Количество непрочитанных сообщений (справа снизу, поверх аватара)
            unread_count = self.get_unread_message_count(user_id)
            if unread_count > 0:
                unread_label = tk.Label(
                    user_frame,
                    text=f"{unread_count}",
                    font=("Helvetica", 10, "bold"),
                    bg="red",
                    fg="white",
                    bd=2,
                    relief="solid",
                )
                unread_label.place(relx=1.0, rely=1.0, x=-10, y=-10, anchor="se")  # Позиционируем справа снизу


# -------------------------------------------------------------------------------------------------------------------------------


# Хешируем пароль "12" через SHA256
VALID_USERNAME = "Skeleton"
VALID_PASSWORD_HASH = hashlib.sha256("12".encode()).hexdigest()

with open(DATA_FILE, "r", encoding="utf-8") as file:
    config = json.load(file)


def get_current_time_kiev():
    kiev_tz = pytz.timezone('Europe/Kiev')
    now = datetime.now(kiev_tz)
    return now.strftime("%H:%M; %d/%m/%Y")


def save_data(data):
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def load_sent_messages():
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    return data.get("sent_messages", {})


def save_sent_messages(sent_messages):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)
    data["sent_messages"] = sent_messages
    with open(DATA_FILE, "w", encoding="utf-8") as file:
        json.dump(data, file, ensure_ascii=False, indent=4)


def load_muted_users_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    muted_users = {}
    for user in data.get("users", []):
        if user.get("mute", False):
            mute_end = user.get("mute_end")
            if mute_end:
                mute_end = datetime.strptime(mute_end, "%H:%M; %d/%m/%Y")
            muted_users[user["id"]] = {
                "first_name": user.get("first_name"),
                "second_name": user.get("second_name"),
                "username": user.get("username"),
                "expiration": mute_end,
                "reason": user.get("reason")
            }
    return muted_users


def load_users_info(json_file=DATA_FILE):
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get("users", [])
    except FileNotFoundError:
        print(f"Помилка: Файл '{json_file}' не знайден.")
        return []
    except json.JSONDecodeError:
        print("Помилка: некорректний формат JSON.")
        return []


def load_chat_id_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    chat_id = data.get("chat_id")
    return chat_id


def load_bottocen_from_file(file_path=DATA_FILE):
    with open(file_path, "r", encoding="utf-8") as file:
        data = json.load(file)

    bot_token = data.get("bot_token")
    return bot_token


def update_data_json(data):
    with open(DATA_FILE, "w") as file:
        json.dump(data, file, indent=4, ensure_ascii=False)


users_info = load_users_info()
muted_users = load_muted_users_from_file()

CREATOR_CHAT_ID = load_chat_id_from_file()
BOTTOCEN = load_bottocen_from_file()
TELEGRAM_API_URL = f"https://api.telegram.org/bot{BOTTOCEN}/"


def load_users(file_path=DATA_FILE):
    """Загружает пользователей и формирует список с доп. инфо."""
    data = load_data(file_path)
    users = []
    for user in data["users"]:
        username = user.get("username", "")
        avatar_url = f"https://t.me/i/userpic/320/{username}.jpg" if username else "https://via.placeholder.com/50"

        mute_status = user.get("mute", False)
        mute_end_date = user.get("mute_end", "None")

        # Определяем статус мута
        if mute_status and mute_end_date != "None":
            status = f"🔴 В муте (до {mute_end_date})"
        else:
            status = "🟢 Размучен"

        users.append({
            "id": user["id"],
            "second_name": user["second_name"],  # Используем second_name
            "username": username,
            "avatar": avatar_url,
            "status": status,
            "rating": user.get("rating", 0),
            "mute_end": mute_end_date
        })
    return users


def get_statistics():
    """Возвращает количество пользователей и среднюю оценку."""
    data = load_data(DATA_FILE)
    total_users = len(data["users"])
    total_score = data.get("total_score", 0)
    num_ratings = data.get("num_of_ratings", 1)
    avg_rating = round(total_score / num_ratings, 1) if num_ratings > 0 else 0
    return total_users, avg_rating


def load_data(filename):
    with open(filename, "r", encoding="utf-8") as f:
        return json.load(f)


def load_chats():
    try:
        with open(CHATS_FILE, "r", encoding="utf-8") as file:
            return json.load(file)
    except FileNotFoundError:
        return {}


# Функция для сохранения чатов в файл
def save_chats(chats):
    with open(CHATS_FILE, "w", encoding="utf-8") as file:
        json.dump(chats, file, ensure_ascii=False, indent=4)


async def start(update: Update, context):
    user = update.message.from_user
    chat_id = update.effective_chat.id

    if chat_id == CREATOR_CHAT_ID:
        await update.message.reply_text("Команда /start недоступна в цій групі.")
        return

    user_found = False
    for u in config["users"]:
        if u["id"] == str(user.id):
            user_found = True
            break

    if not user_found:
        new_user = {
            "id": str(user.id),
            "username": user.username or "Не вказано",
            "first_name": user.first_name or "Не вказано",
            "second_name": user.first_name or "Не вказано",
            "join_date": get_current_time_kiev(),
            "rating": 0,
            "mute": False,
            "mute_end": None,
            "reason": None
        }
        config["users"].append(new_user)
        save_data(config)

    # Загружаем данные чатов
    chats_data = load_chats()

    # Если пользователя нет в чате, добавляем его
    if str(user.id) not in chats_data:
        chats_data[str(user.id)] = {
            "username": user.username or "Не вказано",
            "messages": [
                {
                    "username": "SupportBot",
                    "message": "Привіт! Я ваш бот підтримки. Введіть команду /rate для оцінки бота, /message для написання адміністраторам бота або /help для отримання інформації про команди.",
                    "time_sent": get_current_time_kiev(),
                    "read": True
                }
            ]  # Список сообщений для этого пользователя
        }
        save_chats(chats_data)

    keyboard = [
        ["/start", "/rate"],
        ["/message", "/stopmessage"],
        ["/fromus", "/help"],
    ]

    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        "Привіт! Я ваш бот підтримки. Введіть команду /rate для оцінки бота, /message для написання адміністраторам бота або /help для отримання інформації про команди.",
        reply_markup=reply_markup
    )


async def rate(update: Update, context):
    user_id = update.message.from_user.id

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_rating = None
    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            user_rating = user['rating']
            break

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    rating_text = f"Загальна оцінка: {round(average_rating, 1)}⭐️\nВаш попередній відгук: {user_rating}⭐️"

    keyboard = [
        [InlineKeyboardButton("0.5⭐️", callback_data='0.5'), InlineKeyboardButton("1⭐️", callback_data='1')],
        [InlineKeyboardButton("1.5⭐️", callback_data='1.5'), InlineKeyboardButton("2⭐️", callback_data='2')],
        [InlineKeyboardButton("2.5⭐️", callback_data='2.5'), InlineKeyboardButton("3⭐️", callback_data='3')],
        [InlineKeyboardButton("3.5⭐️", callback_data='3.5'), InlineKeyboardButton("4⭐️", callback_data='4')],
        [InlineKeyboardButton("4.5⭐️", callback_data='4.5'), InlineKeyboardButton("5⭐️", callback_data='5')],
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await update.message.reply_text(f"{rating_text}\nОберіть оцінку:", reply_markup=reply_markup)


async def button_callback(update: Update, context):
    query = update.callback_query
    user_id = query.from_user.id
    new_rating = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    user_found = False
    previous_rating = 0

    for user in data.get("users", []):
        if user.get('id') == str(user_id):
            previous_rating = user.get('rating', 0)
            user['rating'] = new_rating
            user_found = True
            break

    if not user_found:
        new_user = {
            'id': str(user_id),
            'first_name': query.from_user.first_name,
            'username': query.from_user.username,
            'join_date': datetime.now().strftime("%H:%M; %d/%m/%Y"),
            'rating': new_rating,
            'mute': False,
            'mute_end': None,
            'reason': None
        }
        data['users'].append(new_user)

    total_score = data.get("total_score", 0)
    num_of_ratings = data.get("num_of_ratings", 0)

    if previous_rating == 0:
        num_of_ratings += 1
        total_score += new_rating
    else:
        total_score = total_score - previous_rating + new_rating

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=5,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings if num_of_ratings > 0 else 0

    await query.edit_message_text(
        f"Дякуємо за ваш відгук! Ваша оцінка: {new_rating}⭐️\nЗагальна оцінка: {round(average_rating, 1)}⭐️"
    )


async def button(update: Update, context):
    global total_score, num_of_ratings

    query = update.callback_query
    await query.answer()

    selected_rate = float(query.data)

    with open(DATA_FILE, "r", encoding="utf-8") as json_file:
        data = json.load(json_file)

    total_score = data.get("total_score", 0) + selected_rate
    num_of_ratings = data.get("num_of_ratings", 0) + 1

    data["total_score"] = total_score
    data["num_of_ratings"] = num_of_ratings

    with open(DATA_FILE, "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, ensure_ascii=False, indent=4,
                  default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj, datetime) else None)

    average_rating = total_score / num_of_ratings

    user_id = query.from_user.id
    if user_id in users_info:
        users_info[user_id]['rating'] = selected_rate

    await query.edit_message_text(
        f"Дякуємо за ваш відгук! Ваша оцінка: {selected_rate}⭐️\nЗагальна оцінка: {round(average_rating, 1)}⭐️")


async def auto_delete_message(bot, chat_id, message_id, delay):
    await asyncio.sleep(delay)
    await bot.delete_message(chat_id=chat_id, message_id=message_id)


async def message(update: Update, context):
    user_id = update.message.from_user.id
    muted_users = load_muted_users_from_file()

    if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
        reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
        return

    reply = await update.message.reply_text(
        "Введіть ваше повідомлення, і його буде відправлено адміністраторам бота. Введіть /stopmessage, щоб завершити введення повідомлень."
    )

    context.user_data['waiting_for_message'] = True

    await asyncio.create_task(
        auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))


async def stopmessage(update: Update, context):
    if context.user_data.get('waiting_for_message'):
        reply = await update.message.reply_text("Ви завершили введення повідомлень.")
        context.user_data['waiting_for_message'] = False
        await asyncio.create_task(
            auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    else:
        await update.message.reply_text("Ви не в режимі введення повідомлень.")


async def help(update: Update, context):
    if str(update.message.chat.id) == str(CREATOR_CHAT_ID):
        help_text = (
            "Доступні команди в групі:\n"
            "Відповісти на повідомлення бота - Надіслати повідомлення користувачу, який надіслав це повідомлення.\n"
            "/mute <час> <користувач> 'причина' - Замутити користувача на вказаний час.\n"
            "/unmute <користувач> - Розмутити користувача.\n"
            "/mutelist - Показати список замучених користувачів.\n"
            "/alllist - Показати всіх користувачів.\n"
            "/allmessage <повідомлення> - Надіслати повідомлення всім користувачам.\n"
            "/fromus - Інформація про створювача.\n"
            "/help - Показати доступні команди.\n"
            "/info - Показати інформацію про програмістів та адміністраторів.\n"
            "/admin <користувач> - Додати адміністратора.\n"
            "/deleteadmin <користувач> - Видалити адміністратора.\n"
            "/programier <користувач> - Додати програміста.\n"
            "/deleteprogramier <користувач> - Видалити програміста.\n"
            "/get_alllist - Отримати Exel файл з користувачами.\n"
            "/set_alllist - Записати Exel файл з користувачами.\n"
        )
    elif str(update.message.chat.id) == str(-1002358066044):
        help_text = (
            "Доступні команди в групі:\n"
            "/get_alllist - Отримати Exel файл з користувачами.\n"
            "/set_alllist - Записати Exel файл з користувачами.\n"
        )
    else:
        help_text = (
            "Доступні команди в боті:\n"
            "/start - Запустити бота.\n"
            "/rate - Залишити відгук.\n"
            "/message - Почати введення повідомлень адміністраторам.\n"
            "/stopmessage - Завершити введення повідомлень.\n"
            "/fromus - Інформація про створювача.\n"
            "/help - Показати доступні команди.\n"
        )

    await update.message.reply_text(help_text)


async def fromus(update: Update, context):
    await update.message.reply_text(
        "*Skeleton*  Написв бота\nПортфоліо:  ```https://www.linkedin.com/in/artem-k-972a41344/``` \n Телеграм канал з усіма проєктами: ```https://t.me/AboutMyProjects```\n По всім питанням пишіть в цього бота",
        parse_mode="MarkdownV2"
    )


async def info(update: Update, context: CallbackContext):
    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    programmers = data.get("programmers", [])
    admins = data.get("admins", [])

    programmer_list = "\n".join(programmers) if programmers else "Список программистов пуст."
    admin_list = "\n".join(admins) if admins else "Список администраторов пуст."

    await update.message.reply_text(f"Программісти:\n{programmer_list}\n\nАдміністратори:\n{admin_list}")


async def update_website(message_info):
    url = "http://127.0.0.1:5000/update_chat"  # Отправляем локально
    headers = {"Content-Type": "application/json"}

    try:
        response = requests.post(url, json=message_info, headers=headers)
        if response.status_code == 200:
            print("✅ Данные успешно отправлены на сайт")
        else:
            print(f"❌ Ошибка {response.status_code}: {response.text}")
    except Exception as e:
        print(f"Ошибка при отправке данных на сайт: {e}")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    sent_messages = load_sent_messages()
    muted_users = load_muted_users_from_file()
    if context.user_data.get("awaiting_file"):
        if update.message.document:
            document = update.message.document
            file_path = "uploaded_file.xlsx"

            file = await document.get_file()
            await file.download_to_drive(file_path)

            try:
                wb = load_workbook(file_path)

                sheet_all_user = wb["AllUser"]
                sheet_admins = wb["Admins"]
                sheet_programmers = wb["Programmers"]
                sheet_general_info = wb["GeneralInfo"]
                sheet_sent_messages = wb["SentMessages"]

                updated_users = []
                muted_users = {}
                sent_messages = {}

                for row in sheet_all_user.iter_rows(min_row=2, values_only=True):
                    if len(row) < 9:
                        continue

                    user_data = {
                        "id": str(row[0]),
                        "first_name": row[1],
                        "second_name": row[2],
                        "username": row[3],
                        "join_date": row[4].strftime("%H:%M; %d/%m/%Y") if isinstance(row[3], datetime) else str(
                            row[4]),
                        "rating": int(row[5]) if row[5] is not None else 0,
                        "mute": bool(row[6]),
                        "mute_end": row[7].strftime("%H:%M; %d/%m/%Y") if isinstance(row[6], datetime) else str(row[7]),
                        "reason": row[8]
                    }
                    updated_users.append(user_data)

                    if user_data["mute"]:
                        muted_users[user_data["username"]] = True

                for row in sheet_sent_messages.iter_rows(min_row=2, values_only=True):
                    if len(row) < 2 or not row[0] or not row[1]:
                        continue
                    sent_messages[str(row[0])] = row[1]

                admins = [row[0] for row in sheet_admins.iter_rows(min_row=2, values_only=True)]
                programmers = [row[0] for row in sheet_programmers.iter_rows(min_row=2, values_only=True)]

                bot_token = sheet_general_info.cell(row=2, column=1).value or ""
                owner_id = sheet_general_info.cell(row=2, column=2).value or ""
                chat_id = sheet_general_info.cell(row=2, column=3).value or ""
                total_score = float(sheet_general_info.cell(row=2, column=4).value or 0)
                num_of_ratings = int(sheet_general_info.cell(row=2, column=5).value or 0)

                data = {
                    "users": updated_users,
                    "muted_users": muted_users,
                    "admins": admins,
                    "programmers": programmers,
                    "bot_token": bot_token,
                    "owner_id": owner_id,
                    "chat_id": chat_id,
                    "total_score": total_score,
                    "num_of_ratings": num_of_ratings,
                    "sent_messages": sent_messages,
                }

                with open(DATA_FILE, "w", encoding="utf-8") as json_file_obj:
                    json.dump(data, json_file_obj, ensure_ascii=False, indent=4,
                              default=lambda obj: obj.strftime("%H:%M; %d/%m/%Y") if isinstance(obj,
                                                                                                datetime) else None)

                await update.message.reply_text("Файл успешно обработан!")

            except Exception as e:
                await update.message.reply_text(f"Помилка при обробці файла: {e}")

            finally:
                context.user_data["awaiting_file"] = False
        else:
            await update.message.reply_text("Пожалуйста, отправьте Excel-файл.")
    elif (str(update.message.chat.id)) != (str(CREATOR_CHAT_ID)):
        user_id = update.message.from_user.id
        if user_id in muted_users and muted_users[user_id]['expiration'] > datetime.now():
            reply = await update.message.reply_text("Ви в муті й не можете надсилати повідомлення.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

        if context.user_data.get('waiting_for_message'):
            user_name = update.effective_user.first_name
            user_username = update.effective_user.username if update.effective_user.username else "немає імені користувача"
            current_time = get_current_time_kiev()
            user_message = update.message.text if update.message.text else ""

            first_message = f'Повідомлення від **{user_name}**; ```@{user_username}``` \n{current_time}:'
            if user_message:
                first_message += f'\n{user_message}'

            # Загружаем данные чатов из chats.json
            chats_data = load_chats()
            chat_id_str = str(update.message.chat.id)

            # Проверяем, существует ли чат, если нет — создаем
            if chat_id_str not in chats_data or not isinstance(chats_data[chat_id_str], dict):
                print(f"Ошибка: данные для чата {chat_id_str} повреждены, исправляем.")
                chats_data[chat_id_str] = {"username": user_username, "messages": []}

            # Проверяем, что "messages" существует и это список
            if "messages" not in chats_data[chat_id_str] or not isinstance(chats_data[chat_id_str]["messages"], list):
                print(f"Ошибка: messages в чате {chat_id_str} повреждены, исправляем.")
                chats_data[chat_id_str]["messages"] = []

            # Сохраняем информацию о сообщении (без user_id)
            message_info = {
                "username": user_username,
                "message": user_message,
                "time_sent": current_time,
                "read": False
            }

            # Добавляем сообщение в список

            chats_data[chat_id_str]["messages"].append(message_info)
            if chat_id_str in chats_data:
                chats_data[chat_id_str]["newmassages"] = chats_data[chat_id_str].get("newmassages", 0) + 1

            # Сохраняем обновленные данные в chats.json
            save_chats(chats_data)

            # Обновление информации на сайте
            app.update_user_list()  # Тут будет исправленный запрос (см. ниже)====================================================================================

            # Ответ пользователю
            reply = await update.message.reply_text("Ваше повідомлення надіслано адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=5))
    else:
        if update.effective_user.id != context.bot.id:
            if update.message.reply_to_message:
                if update.message.reply_to_message.from_user.id == context.bot.id:
                    original_message_id = str(update.message.reply_to_message.message_id)
                    if original_message_id in sent_messages:
                        original_user_id = sent_messages[original_message_id]
                        reply_text = update.message.text if update.message.text else ""
                        for user in config['users']:
                            if str(user['id']) == str(original_user_id):
                                user_name = user['first_name']
                                break

                        if update.message.photo:
                            photo_file_id = update.message.photo[-1].file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_photo(chat_id=original_user_id, photo=photo_file_id, caption=caption)

                        elif update.message.document:
                            document_file_id = update.message.document.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_document(chat_id=original_user_id, document=document_file_id,
                                                            caption=caption)
                        elif update.message.sticker:
                            sticker_file_id = update.message.sticker.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_sticker(chat_id=original_user_id, sticker=sticker_file_id)

                        elif update.message.voice:
                            voice_file_id = update.message.voice.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_voice(chat_id=original_user_id, voice=voice_file_id, caption=caption)

                        elif update.message.video:
                            video_file_id = update.message.video.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video(chat_id=original_user_id, video=video_file_id, caption=caption)

                        elif update.message.video_note:
                            video_note_file_id = update.message.video_note.file_id
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_video_note(chat_id=original_user_id, video_note=video_note_file_id)
                        else:
                            caption = update.message.caption if update.message.caption else ''
                            await context.bot.send_message(chat_id=original_user_id, text=reply_text)
                        await update.message.reply_text(f"Користувачу {user_name} було надіслано повідомлення")
                        sent_messages[update.message.message_id] = update.message.from_user.id
                        save_sent_messages(sent_messages)


async def mute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    message_text = update.message.text.split()

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return

    mute_time = 300
    reason = "По рішенню адміністратора"
    username = None

    if len(context.args) > 0:
        if context.args[0].isdigit():
            mute_time = int(context.args[0])
            username = context.args[1].lstrip('@') if len(context.args) > 1 else None
        else:
            username = context.args[0].lstrip('@')

    reason_match = re.search(r'["\'](.*?)["\']', update.message.text)
    if reason_match:
        reason = reason_match.group(1)

    if not username:
        await update.message.reply_text("Не вказано користувача для мута.")
        return

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username),
                None)

    if not user:
        await update.message.reply_text(f"Користувач {username} не знайден.")
        return

    if user["id"] == config["owner_id"]:
        await update.message.reply_text("Неможливо замутити власника чату.")
        return

    if user["mute"]:
        await update.message.reply_text(f"Користувач {user['first_name']} вже був замучений.")

    user["mute"] = True
    user["mute_end"] = (datetime.now() + timedelta(seconds=mute_time)).strftime("%H:%M; %d/%m/%Y")
    user["reason"] = reason

    config["muted_users"][username] = True
    save_data(config)

    mute_permissions = ChatPermissions(can_send_messages=False)
    await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"], permissions=mute_permissions)
    await context.bot.send_message(chat_id=user["id"],
                                   text=f"Вас замутили на {str(timedelta(seconds=mute_time))}\nПричина: {reason}")
    await update.message.reply_text(f"Користувач @{user['username']} замучений.")


async def unmute(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /unmute <користувач>")
        return

    username = context.args[0].lstrip('@')

    user = next((u for u in config["users"] if u["username"].lower() == username.lower() or str(u["id"]) == username),
                None)

    if user and user["mute"]:
        user["mute"] = False
        user["mute_end"] = None
        user["reason"] = None

        config["muted_users"].pop(username, None)
        save_data(config)

        mute_permissions = ChatPermissions(can_send_messages=True)
        await context.bot.restrict_chat_member(chat_id=config["chat_id"], user_id=user["id"],
                                               permissions=mute_permissions)
        await update.message.reply_text(f"Користувач @{user['username']} був розмучений.")
    else:
        await update.message.reply_text(f"Користувач {username} не знайден або не був замучений.")


async def admin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("Ця команда доступна тільки програмістам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /admin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        await update.message.reply_text(f"Користувач @{username} вже є администратором.")
    else:
        config["admins"].append(username)
        save_data(config)
        await update.message.reply_text(f"Користувач @{username} додан в список администраторів.")


async def deleteadmin(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if not is_programmer(user):
        await update.message.reply_text("Ця команда доступна тільки програмістам.")
        return

    if len(context.args) < 1:
        await update.message.reply_text("Використовуйте: /deleteadmin @username")
        return

    username = context.args[0].lstrip('@')
    if username in config["admins"]:
        config["admins"].remove(username)
        save_data(config)
        await update.message.reply_text(f"Користувач @{username} видален зі списку администраторів.")
    else:
        await update.message.reply_text(f"Користувач @{username} не знайден.")


async def programier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            new_programmer = context.args[0].replace("@", "")
            if new_programmer not in config["programmers"]:
                config["programmers"].append(new_programmer)
                save_data(config)
                await update.message.reply_text(f"Користувач {new_programmer} додан в список программістів.")
            else:
                await update.message.reply_text(f"Користувач {new_programmer} вже є в списку программистів.")
        else:
            await update.message.reply_text("Використовуйте: /programier @username")
    else:
        await update.message.reply_text("Ця команда доступна лише адміністраторам.")


async def deleteprogramier(update: Update, context: CallbackContext):
    user = update.message.from_user.username
    if is_programmer(user):
        if len(context.args) > 0:
            removed_programmer = context.args[0].replace("@", "")
            if removed_programmer == "ArtemKirss":
                await update.message.reply_text(f"Неможливо видалити {removed_programmer} зі списку программистов.")
            elif removed_programmer in config["programmers"]:
                config["programmers"].remove(removed_programmer)
                save_data(config)
                await update.message.reply_text(f"Користувач {removed_programmer} видален зі списку программистів.")
            else:
                await update.message.reply_text(f"Користувач {removed_programmer} не є программистом.")
        else:
            await update.message.reply_text("Використовуйте: /deleteprogramier @username")
    else:
        await update.message.reply_text("Ця команда доступна лише адміністраторам.")


async def mutelist(update: Update, context):
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "Замучені користувачі:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info.get('mute_end', 'Невідомо')
            reason = mute_info.get('reason', 'Без причини')

            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=int(user_id))
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"

            join_date = mute_info.get('join_date', 'Невідома')
            rating = mute_info.get('rating', 0)
            mute_symbol = "🔇"

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"Залишилось: {expiration}\n"
                f"Причина: {reason}\n"
                f"Дата заходу: {join_date}\n"
                f"Оцінка: {rating}⭐️\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "Немає замучених користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)


async def alllist(update: Update, context: CallbackContext):
    global mute_symbol
    user = update.message.from_user.username
    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна лише адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        data = json.load(file)

    admins = data.get("admins", [])
    programmers = data.get("programmers", [])
    users_info = {user['id']: user for user in data.get("users", [])}
    muted_users = {user['id']: user for user in data.get("users", []) if user.get("mute", False)}

    response = "Користувачі:\n"
    unique_users = {user['id'] for user in data.get("users", [])}

    if unique_users:
        for user_id in unique_users:
            user_data = users_info.get(str(user_id), {})
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"
            join_date = user_data.get('join_date', 'Невідома')
            rating = user_data.get('rating', 0)

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            mute_symbol = "🔇" if str(user_id) in muted_users else "🔊"

            response += f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\nДата заходу: {join_date}\nОцінка: {rating}⭐️\n"
            response += "-------------------------------------------------------------------------\n"
    else:
        response += "Немає користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    response += "==========================================\n"
    response += "\n"
    response += "==========================================\n"
    response += "Замучені користувачі:\n"

    if muted_users:
        for user_id, mute_info in muted_users.items():
            expiration = mute_info['mute_end'] or "Невідомо"
            reason = mute_info.get('reason', "Без причини")
            user_info = await context.bot.get_chat_member(chat_id=CREATOR_CHAT_ID, user_id=user_id)
            user_fullname = user_info.user.first_name or "Невідомий"
            username = user_info.user.username or "Немає імені користувача"
            user_data = users_info.get(str(user_id), {})
            join_date = user_data.get('join_date', 'Невідома')
            rating = user_data.get('rating', 0)

            admins_sumdol = "👨🏻‍💼"
            if username in admins:
                admins_sumdol = "👮🏻‍♂️"
            if username in programmers:
                admins_sumdol = "👨🏻‍💻"

            mute_symbol = "🔇"

            response += (
                f"{admins_sumdol} {mute_symbol} {user_fullname}; @{username} {user_id}\n"
                f"Залишилось: {str(expiration).split('.')[0]}\n"
                f"Причина: {reason}\n"
                f"Дата заходу: {join_date}\n"
                f"Оцінка: {rating}⭐️\n"
                "-------------------------------------------------------------------------\n"
            )
    else:
        response += "Немає замучених користувачів.\n"
        response += "-------------------------------------------------------------------------\n"

    await update.message.reply_text(response)


async def allmessage(update: Update, context):
    user = update.message.from_user.username

    if update.message.chat.id != CREATOR_CHAT_ID:
        if not is_programmer(user) and not is_admin(user):
            reply = await update.message.reply_text("Ця команда доступна тільки адміністраторам бота.")
            await asyncio.create_task(
                auto_delete_message(context.bot, chat_id=reply.chat.id, message_id=reply.message_id, delay=10))
            return

    if not context.args:
        await update.message.reply_text("Будь ласка, укажіть текст повідомлення після команди.")
        return

    message_text = update.message.text.split(' ', 1)[1]

    with open(DATA_FILE, "r", encoding="utf-8") as file:
        config = json.load(file)

    users = config.get("users", [])

    for user_data in users:
        user_id = user_data.get("id")
        if user_id:
            try:
                await context.bot.send_message(chat_id=user_id, text=message_text)
            except Exception as e:
                print(f"Помилка при відправці повідомлення користувачу {user_id}: {e}")

    await update.message.reply_text("Повідомлення відправлено всім користувачам.")


def is_programmer(username):
    return username in config["programmers"]


def is_admin(username):
    return username in config["admins"]


async def get_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])

        users_df = all_users_df[all_users_df["mute"] == False]
        muted_df = all_users_df[all_users_df["mute"] == True]

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime(
                "%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rts2Bot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
            username_cell = row[3]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value),
                               False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:9]:
                cell.fill = fill_color

        workbook.save(excel_file)

        await update.message.reply_document(document=open(excel_file, "rb"))

    except Exception as e:
        await update.message.reply_text(f"Error: {e}")


async def set_alllist(update: Update, context: CallbackContext) -> None:
    user = update.message.from_user.username

    if not is_programmer(user) and not is_admin(user):
        await update.message.reply_text("Ця команда доступна тільки адміністраторам.")
        return
    await update.message.reply_text("Будь ласка пришліть Excel file з данними.")
    context.user_data["awaiting_file"] = True


"""async def set_default_commands(application):
    commands = [
        BotCommand("start", "Запустити бота"),
        BotCommand("rate", "Залишити відгук"),
        BotCommand("message", "Почати введення повідомлень адміністраторам"),
        BotCommand("stopmessage", "Завершити введення повідомлень"),
        BotCommand("fromus", "Інформація про створювача"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeDefault())

async def set_creator_commands(application):
    commands = [
        BotCommand("mutelist", "Показати список замучених користувачів"),
        BotCommand("alllist", "Показати всіх користувачів"),
        BotCommand("fromus", "Інформація про створювача"),
        BotCommand("help", "Показати доступні команди"),
        BotCommand("info", "Показати інформацію про програмістів та адміністраторів"),
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
        print(CREATOR_CHAT_ID)
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=CREATOR_CHAT_ID))

async def set_save_commands(application):
    commands = [
        BotCommand("get_alllist", "Отримати Exel файл з користувачами"),
        BotCommand("set_alllist", "Записати Exel файл з користувачами"),
        BotCommand("help", "Показати доступні команди"),
    ]
    await application.bot.set_my_commands(commands, scope=BotCommandScopeChat(chat_id=-1002310142084))"""


async def send_user_list():
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as file:
            data = json.load(file)

        all_users_df = pd.DataFrame(data["users"])
        print(all_users_df)
        users_df = all_users_df[all_users_df["mute"] == False]
        print(users_df)
        muted_df = all_users_df[all_users_df["mute"] == True]
        print(muted_df)

        muted_df.loc[:, "mute_end"] = muted_df["mute_end"].apply(
            lambda x: datetime.strptime(x.replace(";", " "), "%H:%M %d/%m/%Y").strftime(
                "%H:%M; %d/%m/%Y") if isinstance(x, str) else ""
        )

        admins_df = pd.DataFrame(data.get("admins", []), columns=["Admins"])
        programmers_df = pd.DataFrame(data.get("programmers", []), columns=["Programmers"])
        general_info_df = pd.DataFrame(
            [{
                "bot_token": data.get("bot_token"),
                "owner_id": data.get("owner_id"),
                "chat_id": data.get("chat_id"),
                "total_score": data.get("total_score"),
                "num_of_ratings": data.get("num_of_ratings")
            }]
        )
        sent_messages_df = pd.DataFrame(data.get("sent_messages", {}).items(), columns=["MessageID", "UserID"])
        muted_users_df = pd.DataFrame(data.get("muted_users", {}).items(), columns=["UserID", "Details"])

        excel_file = "Supp0rts2Bot_all_users.xlsx"
        with pd.ExcelWriter(excel_file) as writer:
            users_df.to_excel(writer, index=False, sheet_name="Users")
            muted_df.to_excel(writer, index=False, sheet_name="Muted")
            all_users_df.to_excel(writer, index=False, sheet_name="AllUser")
            admins_df.to_excel(writer, index=False, sheet_name="Admins")
            programmers_df.to_excel(writer, index=False, sheet_name="Programmers")
            general_info_df.to_excel(writer, index=False, sheet_name="GeneralInfo")
            sent_messages_df.to_excel(writer, index=False, sheet_name="SentMessages")
            muted_users_df.to_excel(writer, index=False, sheet_name="MutedUsers")

        workbook = load_workbook(excel_file)
        sheet = workbook["AllUser"]

        yellow_fill = PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
        red_fill = PatternFill(start_color="b40a0a", end_color="b40a0a", fill_type="solid")

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8):
            username_cell = row[2]
            mute_status = next((user['mute'] for user in data["users"] if user["username"] == username_cell.value),
                               False)

            fill_color = red_fill if mute_status else yellow_fill

            for cell in row[:8]:
                cell.fill = fill_color

        workbook.save(excel_file)

        bot = Bot(token=BOTTOCEN)
        await bot.send_document(chat_id=-1002358066044, document=open(excel_file, "rb"))

    except Exception as e:
        bot = Bot(token=BOTTOCEN)
        await bot.send_message(chat_id=-1002358066044, text=f"Ошибка при создании отчета: {e}")


async def main():
    application = Application.builder().token("7677888606:AAHMm3aSt84ZQkJ0wrlH4__St3lW36-TL8g").build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("rate", rate))
    application.add_handler(CommandHandler("message", message))
    application.add_handler(CommandHandler("stopmessage", stopmessage))
    application.add_handler(CommandHandler("fromus", fromus))
    application.add_handler(CommandHandler("help", help))
    application.add_handler(CommandHandler("mute", mute))
    application.add_handler(CommandHandler("unmute", unmute))
    application.add_handler(CommandHandler("mutelist", mutelist))
    application.add_handler(CommandHandler("alllist", alllist))
    application.add_handler(CommandHandler("allmessage", allmessage))
    application.add_handler(CommandHandler("admin", admin))
    application.add_handler(CommandHandler("deleteadmin", deleteadmin))
    application.add_handler(CommandHandler("programier", programier))
    application.add_handler(CommandHandler("deleteprogramier", deleteprogramier))
    application.add_handler(CommandHandler("info", info))
    application.add_handler(CommandHandler("get_alllist", get_alllist))
    application.add_handler(CommandHandler("set_alllist", set_alllist))

    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.ALL, handle_message))

    """await set_default_commands(application)
    await set_creator_commands(application)
    await set_save_commands(application)"""

    scheduler = AsyncIOScheduler(timezone=pytz.timezone("Europe/Kyiv"))
    scheduler.add_job(send_user_list, "cron", hour=0, minute=0)
    scheduler.start()

    application.run_polling()


def run_flask():
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)


def run_bot():
    """Запускает Telegram-бота в отдельном потоке."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(main())


if __name__ == "__main__":
    threading.Thread(target=run_flask, daemon=True).start()
    threading.Thread(target=run_bot, daemon=True).start()

    root = tk.Tk()
    app = ChatApp(root)
    root.mainloop()


